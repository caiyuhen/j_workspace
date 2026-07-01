"""
Excel (.xlsx) 文档导出模块
Spreadsheet Exporter Module

支持:
- Meta 分析结果导出
- 经费预算表导出
- 期刊数据库导出
- 生存分析数据导出
"""

import os
from typing import Dict, List, Any, Optional
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, Reference
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class BaseSpreadsheetExporter:
    """Excel 导出基类"""

    def __init__(self):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required. Install: pip install openpyxl")
        self.wb = Workbook()
        self.ws = self.wb.active

    def _setup_header_style(self, row: int, cols: int):
        """设置表头样式"""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col in range(1, cols + 1):
            cell = self.ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

    def _apply_zebra_striping(self, start_row: int, end_row: int, cols: int):
        """应用斑马纹"""
        fill_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        fill_odd = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        for row in range(start_row, end_row + 1):
            fill = fill_even if (row - start_row) % 2 == 0 else fill_odd
            for col in range(1, cols + 1):
                self.ws.cell(row=row, column=col).fill = fill

    def _auto_column_width(self, min_width: int = 10, max_width: int = 50):
        """自动调整列宽"""
        for column in self.ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max(min_width, max_length + 2), max_width)
            self.ws.column_dimensions[column_letter].width = adjusted_width

    def save(self, file_path: str):
        """保存工作簿"""
        os.makedirs(os.path.dirname(file_path) or '.', exist_ok=True)
        self.wb.save(file_path)
        return file_path


class MetaAnalysisExporter(BaseSpreadsheetExporter):
    """Meta 分析结果导出器"""

    def export_meta_analysis(self, result: Dict[str, Any], file_path: str) -> str:
        """
        将 Meta 分析结果导出为 Excel

        Args:
            result: MetaAnalysisResult 的字典表示
            file_path: 输出路径
        """
        self.ws.title = "Meta Analysis"

        # 标题
        self.ws.merge_cells("A1:F1")
        title_cell = self.ws["A1"]
        title_cell.value = "Meta-Analysis Results"
        title_cell.font = Font(bold=True, size=16, color="1F4E78")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        self.ws.row_dimensions[1].height = 30

        # 研究数据
        row = 3
        self.ws.cell(row=row, column=1, value="Study Data")
        self.ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        headers = ["Study", "Group A Events", "Group A Total", "Group B Events", "Group B Total", "Effect Size"]
        for i, h in enumerate(headers, 1):
            self.ws.cell(row=row, column=i, value=h)
        self._setup_header_style(row, len(headers))

        studies = result.get("studies", [])
        start_data_row = row + 1
        for i, study in enumerate(studies):
            r = row + 1 + i
            self.ws.cell(row=r, column=1, value=study.get("name", f"Study {i+1}"))
            self.ws.cell(row=r, column=2, value=study.get("a_events", 0))
            self.ws.cell(row=r, column=3, value=study.get("a_total", 0))
            self.ws.cell(row=r, column=4, value=study.get("b_events", 0))
            self.ws.cell(row=r, column=5, value=study.get("b_total", 0))
            self.ws.cell(row=r, column=6, value=round(study.get("effect_size", 0), 3))

        end_data_row = row + len(studies)
        self._apply_zebra_striping(start_data_row, end_data_row, len(headers))

        # 汇总统计
        row = end_data_row + 2
        self.ws.cell(row=row, column=1, value="Summary Statistics")
        self.ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        summary_items = [
            ("Pooled Effect Size", result.get("pooled_effect", "N/A")),
            ("95% CI Lower", result.get("ci_lower", "N/A")),
            ("95% CI Upper", result.get("ci_upper", "N/A")),
            ("Heterogeneity I²", result.get("i_squared", "N/A")),
            ("Q Statistic", result.get("q_statistic", "N/A")),
            ("P-value", result.get("p_value", "N/A")),
            ("Model", result.get("model", "Random")),
        ]

        for label, value in summary_items:
            self.ws.cell(row=row, column=1, value=label)
            self.ws.cell(row=row, column=1).font = Font(bold=True)
            self.ws.cell(row=row, column=2, value=value)
            row += 1

        self._auto_column_width()
        return self.save(file_path)


class BudgetExporter(BaseSpreadsheetExporter):
    """经费预算表导出器"""

    def export_budget(self, budget_data: Dict[str, Any], file_path: str) -> str:
        """
        将经费预算导出为 Excel

        Args:
            budget_data: 预算数据，包含 items 列表和 total
            file_path: 输出路径
        """
        self.ws.title = "Budget"

        # 标题
        self.ws.merge_cells("A1:E1")
        title = budget_data.get("title", "Research Budget")
        self.ws["A1"] = title
        self.ws["A1"].font = Font(bold=True, size=16, color="1F4E78")
        self.ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        self.ws.row_dimensions[1].height = 30

        # 表头
        row = 3
        headers = ["No.", "Budget Item", "Amount (10K CNY)", "Percentage", "Notes"]
        for i, h in enumerate(headers, 1):
            self.ws.cell(row=row, column=i, value=h)
        self._setup_header_style(row, len(headers))

        # 数据
        items = budget_data.get("items", [])
        total = budget_data.get("total", 0)
        start_data_row = row + 1

        for i, item in enumerate(items, 1):
            r = row + i
            amount = item.get("amount", 0)
            pct = f"{amount / total * 100:.1f}%" if total > 0 else "0%"
            self.ws.cell(row=r, column=1, value=i)
            self.ws.cell(row=r, column=2, value=item.get("name", ""))
            self.ws.cell(row=r, column=3, value=round(amount, 2))
            self.ws.cell(row=r, column=4, value=pct)
            self.ws.cell(row=r, column=5, value=item.get("notes", ""))

        end_data_row = row + len(items)
        self._apply_zebra_striping(start_data_row, end_data_row, len(headers))

        # 合计行
        total_row = end_data_row + 1
        self.ws.cell(row=total_row, column=2, value="Total")
        self.ws.cell(row=total_row, column=2).font = Font(bold=True)
        self.ws.cell(row=total_row, column=3, value=round(total, 2))
        self.ws.cell(row=total_row, column=3).font = Font(bold=True)
        self.ws.cell(row=total_row, column=4, value="100%")
        self.ws.cell(row=total_row, column=4).font = Font(bold=True)

        # 添加边框
        thin_border = Border(
            left=Side(style="thin", color="D9DEE7"),
            right=Side(style="thin", color="D9DEE7"),
            top=Side(style="thin", color="D9DEE7"),
            bottom=Side(style="thin", color="D9DEE7"),
        )
        for r in range(start_data_row, total_row + 1):
            for c in range(1, len(headers) + 1):
                self.ws.cell(row=r, column=c).border = thin_border

        self._auto_column_width()
        return self.save(file_path)


class JournalDatabaseExporter(BaseSpreadsheetExporter):
    """期刊数据库导出器"""

    def export_journals(self, journals: List[Dict[str, Any]],
                        file_path: str) -> str:
        """
        将期刊数据库导出为 Excel

        Args:
            journals: 期刊列表
            file_path: 输出路径
        """
        self.ws.title = "Journals"

        # 标题
        self.ws.merge_cells("A1:H1")
        self.ws["A1"] = "Medical Journal Database"
        self.ws["A1"].font = Font(bold=True, size=16, color="1F4E78")
        self.ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        self.ws.row_dimensions[1].height = 30

        # 表头
        row = 3
        headers = ["No.", "Journal Name", "Impact Factor", "JCR Quartile",
                   "CAS Quartile", "Field", "OA Policy", "Review Period"]
        for i, h in enumerate(headers, 1):
            self.ws.cell(row=row, column=i, value=h)
        self._setup_header_style(row, len(headers))

        start_data_row = row + 1
        for i, journal in enumerate(journals, 1):
            r = row + i
            self.ws.cell(row=r, column=1, value=i)
            self.ws.cell(row=r, column=2, value=journal.get("name", ""))
            self.ws.cell(row=r, column=3, value=journal.get("impact_factor", ""))
            self.ws.cell(row=r, column=4, value=journal.get("jcr_quartile", ""))
            self.ws.cell(row=r, column=5, value=journal.get("cas_quartile", ""))
            self.ws.cell(row=r, column=6, value=journal.get("field", ""))
            self.ws.cell(row=r, column=7, value=journal.get("oa_policy", ""))
            self.ws.cell(row=r, column=8, value=journal.get("review_period", ""))

        end_data_row = row + len(journals)
        self._apply_zebra_striping(start_data_row, end_data_row, len(headers))
        self._auto_column_width()
        return self.save(file_path)


class SurvivalDataExporter(BaseSpreadsheetExporter):
    """生存分析数据导出器"""

    def export_survival_data(self, records: List[Dict[str, Any]],
                             file_path: str) -> str:
        """
        将生存分析数据导出为 Excel

        Args:
            records: 生存记录列表
            file_path: 输出路径
        """
        self.ws.title = "Survival Data"

        # 标题
        self.ws.merge_cells("A1:F1")
        self.ws["A1"] = "Survival Analysis Data"
        self.ws["A1"].font = Font(bold=True, size=16, color="1F4E78")
        self.ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        self.ws.row_dimensions[1].height = 30

        # 表头
        row = 3
        headers = ["Patient ID", "Time (months)", "Event (0=censored, 1=event)",
                   "Group", "Age", "Stage"]
        for i, h in enumerate(headers, 1):
            self.ws.cell(row=row, column=i, value=h)
        self._setup_header_style(row, len(headers))

        start_data_row = row + 1
        for i, record in enumerate(records):
            r = row + 1 + i
            self.ws.cell(row=r, column=1, value=record.get("patient_id", ""))
            self.ws.cell(row=r, column=2, value=record.get("time", 0))
            self.ws.cell(row=r, column=3, value=record.get("event", 0))
            self.ws.cell(row=r, column=4, value=record.get("group", ""))
            self.ws.cell(row=r, column=5, value=record.get("age", ""))
            self.ws.cell(row=r, column=6, value=record.get("stage", ""))

        end_data_row = row + len(records)
        self._apply_zebra_striping(start_data_row, end_data_row, len(headers))
        self._auto_column_width()
        return self.save(file_path)
